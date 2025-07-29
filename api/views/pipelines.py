"""
Pipeline API views
"""
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.utils import extend_schema, OpenApiParameter
from django.utils import timezone
from django.http import HttpResponse
from django.db.models import Count
import csv
import io
import json

from pipelines.models import Pipeline, Field, Record
from api.serializers import (
    PipelineSerializer, PipelineListSerializer, FieldSerializer,
    RecordSerializer, DynamicRecordSerializer
)
from api.filters import PipelineFilter
from api.permissions import PipelinePermission
from authentication.permissions import AsyncPermissionManager as PermissionManager


class PipelineViewSet(viewsets.ModelViewSet):
    """
    Pipeline management API with comprehensive CRUD operations
    """
    permission_classes = [permissions.IsAuthenticated, PipelinePermission]
    filter_backends = [DjangoFilterBackend]
    filterset_class = PipelineFilter
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created_at', 'updated_at', 'record_count']
    ordering = ['-created_at']
    
    def get_queryset(self):
        """Get filtered queryset based on user permissions"""
        user = self.request.user
        queryset = Pipeline.objects.filter(is_active=True).annotate(
            record_count=Count('records', filter={'records__is_deleted': False})
        ).select_related('created_by').prefetch_related('fields')
        
        # Apply permission filtering
        permission_manager = PermissionManager(user)
        if not permission_manager.has_permission('action', 'pipelines', 'read_all'):
            # Filter to only accessible pipelines
            accessible_pipeline_ids = self._get_accessible_pipeline_ids(user)
            queryset = queryset.filter(id__in=accessible_pipeline_ids)
        
        return queryset
    
    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'list':
            return PipelineListSerializer
        return PipelineSerializer
    
    def _get_accessible_pipeline_ids(self, user):
        """Get pipeline IDs that user has access to"""
        permission_manager = PermissionManager(user)
        accessible_ids = []
        
        # Check each pipeline for read access
        for pipeline in Pipeline.objects.filter(is_active=True):
            if permission_manager.has_permission('action', 'pipelines', 'read', str(pipeline.id)):
                accessible_ids.append(pipeline.id)
        
        return accessible_ids
    
    @extend_schema(
        summary="Get pipeline analytics",
        description="Retrieve comprehensive analytics data for a specific pipeline",
        responses={200: {
            "type": "object",
            "properties": {
                "record_count": {"type": "integer"},
                "records_created_today": {"type": "integer"},
                "records_updated_today": {"type": "integer"},
                "field_count": {"type": "integer"},
                "relationship_count": {"type": "integer"},
                "active_users": {"type": "integer"},
                "recent_activity": {"type": "array"}
            }
        }}
    )
    @action(detail=True, methods=['get'])
    def analytics(self, request, pk=None):
        """Get analytics data for a pipeline"""
        pipeline = self.get_object()
        
        # Calculate analytics
        today = timezone.now().date()
        analytics_data = {
            'record_count': pipeline.record_count,
            'records_created_today': pipeline.records.filter(
                created_at__date=today,
                is_deleted=False
            ).count(),
            'records_updated_today': pipeline.records.filter(
                updated_at__date=today,
                is_deleted=False
            ).count(),
            'field_count': pipeline.fields.count(),
            'relationship_count': (
                pipeline.outgoing_relationships.filter(is_deleted=False).count() + 
                pipeline.incoming_relationships.filter(is_deleted=False).count()
            ),
            'active_users': self._get_active_users_count(pipeline),
            'recent_activity': self._get_recent_activity(pipeline),
            'status_distribution': self._get_status_distribution(pipeline),
            'creation_trends': self._get_creation_trends(pipeline)
        }
        
        return Response(analytics_data)
    
    @extend_schema(
        summary="Export pipeline data",
        description="Export pipeline records in various formats",
        parameters=[
            OpenApiParameter(
                'format', 
                str, 
                description='Export format: csv, json, xlsx',
                enum=['csv', 'json', 'xlsx']
            ),
            OpenApiParameter(
                'include_deleted', 
                bool, 
                description='Include deleted records'
            )
        ]
    )
    @action(detail=True, methods=['get'])
    def export(self, request, pk=None):
        """Export pipeline data in various formats"""
        pipeline = self.get_object()
        export_format = request.query_params.get('format', 'json')
        include_deleted = request.query_params.get('include_deleted', 'false').lower() == 'true'
        
        # Get records
        records_query = pipeline.records.all()
        if not include_deleted:
            records_query = records_query.filter(is_deleted=False)
        
        records = records_query.order_by('-created_at')
        
        if export_format == 'csv':
            return self._export_csv(pipeline, records)
        elif export_format == 'xlsx':
            return self._export_xlsx(pipeline, records)
        else:
            return self._export_json(pipeline, records)
    
    @extend_schema(
        summary="Clone pipeline",
        description="Create a copy of the pipeline with all its fields",
        request={
            "type": "object", 
            "properties": {
                "name": {"type": "string", "description": "Name for the cloned pipeline"}
            }
        }
    )
    @action(detail=True, methods=['post'])
    def clone(self, request, pk=None):
        """Clone a pipeline with all its fields"""
        source_pipeline = self.get_object()
        new_name = request.data.get('name', f"{source_pipeline.name} (Copy)")
        
        # Clone pipeline
        cloned_pipeline = Pipeline.objects.create(
            name=new_name,
            description=f"Cloned from: {source_pipeline.description}",
            pipeline_type=source_pipeline.pipeline_type,
            icon=source_pipeline.icon,
            color=source_pipeline.color,
            settings=source_pipeline.settings.copy() if source_pipeline.settings else {},
            created_by=request.user
        )
        
        # Clone fields
        for field in source_pipeline.fields.all():
            Field.objects.create(
                pipeline=cloned_pipeline,
                name=field.name,
                field_type=field.field_type,
                field_config=field.field_config.copy() if field.field_config else {},
                is_required=field.is_required,
                is_visible_in_list=field.is_visible_in_list,
                display_order=field.display_order,
                help_text=field.help_text,
                validation_rules=field.validation_rules.copy() if field.validation_rules else {},
                created_by=request.user
            )
        
        serializer = self.get_serializer(cloned_pipeline)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    @extend_schema(
        summary="Pipeline schema",
        description="Get the field schema for this pipeline"
    )
    @action(detail=True, methods=['get'])
    def schema(self, request, pk=None):
        """Get pipeline field schema"""
        pipeline = self.get_object()
        
        schema = {
            'pipeline_id': pipeline.id,
            'pipeline_name': pipeline.name,
            'fields': []
        }
        
        for field in pipeline.fields.all().order_by('display_order'):
            field_schema = {
                'name': field.name,
                'slug': field.slug,
                'type': field.field_type,
                'required': field.is_required,
                'visible_in_list': field.is_visible_in_list,
                'help_text': field.help_text,
                'config': field.field_config,
                'validation_rules': field.validation_rules
            }
            schema['fields'].append(field_schema)
        
        return Response(schema)
    
    def _get_active_users_count(self, pipeline):
        """Get count of users who have interacted with this pipeline recently"""
        from django.utils import timezone
        from datetime import timedelta
        
        week_ago = timezone.now() - timedelta(days=7)
        return pipeline.records.filter(
            updated_at__gte=week_ago,
            is_deleted=False
        ).values('updated_by').distinct().count()
    
    def _get_recent_activity(self, pipeline):
        """Get recent activity for the pipeline"""
        recent_records = pipeline.records.filter(
            is_deleted=False
        ).order_by('-updated_at')[:10]
        
        activity = []
        for record in recent_records:
            activity.append({
                'record_id': record.id,
                'record_title': record.title,
                'action': 'updated',
                'user': record.updated_by.email if record.updated_by else None,
                'timestamp': record.updated_at
            })
        
        return activity
    
    def _get_status_distribution(self, pipeline):
        """Get distribution of record statuses"""
        from django.db.models import Count
        
        distribution = pipeline.records.filter(
            is_deleted=False
        ).values('status').annotate(
            count=Count('id')
        ).order_by('status')
        
        return {item['status']: item['count'] for item in distribution}
    
    def _get_creation_trends(self, pipeline):
        """Get record creation trends over the last 30 days"""
        from django.utils import timezone
        from datetime import timedelta
        from django.db.models import Count
        from django.db.models.functions import TruncDate
        
        thirty_days_ago = timezone.now() - timedelta(days=30)
        
        trends = pipeline.records.filter(
            created_at__gte=thirty_days_ago,
            is_deleted=False
        ).extra({
            'date': 'date(created_at)'
        }).values('date').annotate(
            count=Count('id')
        ).order_by('date')
        
        return [
            {
                'date': item['date'],
                'count': item['count']
            }
            for item in trends
        ]
    
    def _export_csv(self, pipeline, records):
        """Export records as CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{pipeline.slug}_export.csv"'
        
        writer = csv.writer(response)
        
        # Header row
        headers = ['ID', 'Title', 'Status', 'Created At', 'Updated At']
        for field in pipeline.fields.all().order_by('display_order'):
            headers.append(field.name)
        writer.writerow(headers)
        
        # Data rows
        for record in records:
            row = [
                record.id,
                record.title,
                record.status,
                record.created_at.isoformat(),
                record.updated_at.isoformat()
            ]
            
            for field in pipeline.fields.all().order_by('display_order'):
                value = record.data.get(field.slug, '')
                if isinstance(value, (dict, list)):
                    value = json.dumps(value)
                row.append(value)
            
            writer.writerow(row)
        
        return response
    
    def _export_json(self, pipeline, records):
        """Export records as JSON"""
        data = {
            'pipeline': {
                'id': pipeline.id,
                'name': pipeline.name,
                'exported_at': timezone.now().isoformat()
            },
            'records': []
        }
        
        for record in records:
            record_data = {
                'id': record.id,
                'title': record.title,
                'status': record.status,
                'data': record.data,
                'created_at': record.created_at.isoformat(),
                'updated_at': record.updated_at.isoformat()
            }
            data['records'].append(record_data)
        
        response = HttpResponse(
            json.dumps(data, indent=2),
            content_type='application/json'
        )
        response['Content-Disposition'] = f'attachment; filename="{pipeline.slug}_export.json"'
        
        return response
    
    def _export_xlsx(self, pipeline, records):
        """Export records as Excel file"""
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            return Response(
                {'error': 'openpyxl not installed'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = pipeline.name[:31]  # Excel sheet name limit
        
        # Headers
        headers = ['ID', 'Title', 'Status', 'Created At', 'Updated At']
        for field in pipeline.fields.all().order_by('display_order'):
            headers.append(field.name)
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data
        for row_num, record in enumerate(records, 2):
            ws.cell(row=row_num, column=1, value=record.id)
            ws.cell(row=row_num, column=2, value=record.title)
            ws.cell(row=row_num, column=3, value=record.status)
            ws.cell(row=row_num, column=4, value=record.created_at)
            ws.cell(row=row_num, column=5, value=record.updated_at)
            
            col = 6
            for field in pipeline.fields.all().order_by('display_order'):
                value = record.data.get(field.slug, '')
                if isinstance(value, (dict, list)):
                    value = json.dumps(value)
                ws.cell(row=row_num, column=col, value=value)
                col += 1
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{pipeline.slug}_export.xlsx"'
        
        return response


class FieldViewSet(viewsets.ModelViewSet):
    """
    Pipeline field management API
    """
    serializer_class = FieldSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['field_type', 'is_required', 'is_visible_in_list']
    ordering = ['display_order', 'name']
    
    def get_queryset(self):
        """Get fields for a specific pipeline"""
        pipeline_pk = self.kwargs.get('pipeline_pk')
        if pipeline_pk:
            return Field.objects.filter(
                pipeline_id=pipeline_pk
            ).select_related('pipeline', 'created_by')
        return Field.objects.none()
    
    def perform_create(self, serializer):
        """Set pipeline and user when creating field"""
        pipeline_pk = self.kwargs.get('pipeline_pk')
        pipeline = Pipeline.objects.get(id=pipeline_pk)
        serializer.save(
            pipeline=pipeline,
            created_by=self.request.user
        )
    
    @extend_schema(
        summary="Reorder fields",
        description="Update the display order of multiple fields",
        request={
            "type": "object",
            "properties": {
                "field_orders": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "id": {"type": "integer"},
                            "display_order": {"type": "integer"}
                        }
                    }
                }
            }
        }
    )
    @action(detail=False, methods=['post'])
    def reorder(self, request, pipeline_pk=None):
        """Reorder pipeline fields"""
        field_orders = request.data.get('field_orders', [])
        
        updated_count = 0
        for item in field_orders:
            field_id = item.get('id')
            new_order = item.get('display_order')
            
            if field_id and new_order is not None:
                Field.objects.filter(
                    id=field_id,
                    pipeline_id=pipeline_pk
                ).update(display_order=new_order)
                updated_count += 1
        
        return Response({
            'updated_count': updated_count,
            'message': f'Updated display order for {updated_count} fields'
        })