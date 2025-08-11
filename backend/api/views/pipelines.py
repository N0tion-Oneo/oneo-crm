"""
Pipeline API views
"""
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.utils import extend_schema, OpenApiParameter
from django.utils import timezone
from django.http import HttpResponse
from django.db.models import Count
from django.db import models
import csv
import io
import json

from pipelines.models import Pipeline, Field, Record
from api.serializers import (
    PipelineSerializer, PipelineListSerializer, FieldSerializer,
    RecordSerializer, DynamicRecordSerializer
)
from pipelines.serializers import FieldManagementActionSerializer
from api.filters import PipelineFilter
from api.permissions import PipelinePermission
from authentication.permissions import SyncPermissionManager as PermissionManager
from rest_framework.exceptions import PermissionDenied, NotFound


class PipelineViewSet(viewsets.ModelViewSet):
    """
    Pipeline management API with comprehensive CRUD operations
    """
    permission_classes = [PipelinePermission]
    filter_backends = [DjangoFilterBackend]
    filterset_class = PipelineFilter
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created_at', 'updated_at', 'record_count']
    ordering = ['-created_at']
    
    def get_queryset(self):
        """Get filtered queryset based on user permissions"""
        user = self.request.user
        queryset = Pipeline.objects.filter(is_active=True).annotate(
            active_record_count=Count('records', filter=models.Q(records__is_deleted=False))
        ).select_related('created_by').prefetch_related('fields')
        
        # Apply permission filtering
        permission_manager = PermissionManager(user)
        if not permission_manager.has_permission('action', 'pipelines', 'read_all'):
            # Filter to only accessible pipelines
            accessible_pipeline_ids = self._get_accessible_pipeline_ids(user, permission_manager)
            queryset = queryset.filter(id__in=accessible_pipeline_ids)
        
        return queryset
    
    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'list':
            return PipelineListSerializer
        return PipelineSerializer
    
    def _get_accessible_pipeline_ids(self, user, permission_manager):
        """Get pipeline IDs that user has access to via dynamic permissions"""
        from authentication.models import UserTypePipelinePermission
        
        # Only check dynamic permissions - ignore static permissions and access levels
        # This ensures dynamic permissions are properly enforced
        user_type_permissions = UserTypePipelinePermission.objects.filter(
            user_type=user.user_type
        ).values_list('pipeline_id', flat=True)
        
        return list(user_type_permissions)
    
    @extend_schema(
        summary="Get pipeline analytics",
        description="Retrieve comprehensive analytics data for a specific pipeline",
        responses={200: {
            "type": "object",
            "properties": {
                "record_count": {"type": "integer"},
                "records_created_today": {"type": "integer"},
                "records_updated_today": {"type": "integer"},
                "field_count": {"type": "integer"},
                "relationship_count": {"type": "integer"},
                "active_users": {"type": "integer"},
                "recent_activity": {"type": "array"}
            }
        }}
    )
    @action(detail=True, methods=['get'])
    def analytics(self, request, pk=None):
        """Get analytics data for a pipeline"""
        pipeline = self.get_object()
        
        # Calculate analytics
        today = timezone.now().date()
        analytics_data = {
            'record_count': pipeline.record_count,
            'records_created_today': pipeline.records.filter(
                created_at__date=today,
                is_deleted=False
            ).count(),
            'records_updated_today': pipeline.records.filter(
                updated_at__date=today,
                is_deleted=False
            ).count(),
            'field_count': pipeline.fields.count(),
            'relationship_count': (
                pipeline.outgoing_relationships.filter(is_deleted=False).count() + 
                pipeline.incoming_relationships.filter(is_deleted=False).count()
            ),
            'active_users': self._get_active_users_count(pipeline),
            'recent_activity': self._get_recent_activity(pipeline),
            'status_distribution': self._get_status_distribution(pipeline),
            'creation_trends': self._get_creation_trends(pipeline)
        }
        
        return Response(analytics_data)
    
    @extend_schema(
        summary="Export pipeline data",
        description="Export pipeline records in various formats",
        parameters=[
            OpenApiParameter(
                'format', 
                str, 
                description='Export format: csv, json, xlsx',
                enum=['csv', 'json', 'xlsx']
            ),
            OpenApiParameter(
                'include_deleted', 
                bool, 
                description='Include deleted records'
            )
        ]
    )
    @action(detail=True, methods=['get'])
    def export(self, request, pk=None):
        """Export pipeline data in various formats"""
        pipeline = self.get_object()
        export_format = request.query_params.get('format', 'json')
        include_deleted = request.query_params.get('include_deleted', 'false').lower() == 'true'
        
        # Check if user has BOTH pipeline read AND field read permission for including field data
        from authentication.permissions import SyncPermissionManager as PermissionManager
        permission_manager = PermissionManager(request.user)
        pipeline_access = permission_manager.has_permission('action', 'pipelines', 'read', str(pipeline.id))
        field_access = permission_manager.has_permission('action', 'fields', 'read', str(pipeline.id))
        has_field_access = pipeline_access and field_access
        
        # Get records
        records_query = pipeline.records.all()
        if not include_deleted:
            records_query = records_query.filter(is_deleted=False)
        
        records = records_query.order_by('-created_at')
        
        if export_format == 'csv':
            return self._export_csv(pipeline, records)
        elif export_format == 'xlsx':
            return self._export_xlsx(pipeline, records)
        else:
            return self._export_json(pipeline, records)
    
    @extend_schema(
        summary="Clone pipeline",
        description="Create a copy of the pipeline with all its fields",
        request={
            "type": "object", 
            "properties": {
                "name": {"type": "string", "description": "Name for the cloned pipeline"}
            }
        }
    )
    @action(detail=True, methods=['post'])
    def clone(self, request, pk=None):
        """Clone a pipeline with all its fields"""
        source_pipeline = self.get_object()
        new_name = request.data.get('name', f"{source_pipeline.name} (Copy)")
        
        # Clone pipeline
        cloned_pipeline = Pipeline.objects.create(
            name=new_name,
            description=f"Cloned from: {source_pipeline.description}",
            pipeline_type=source_pipeline.pipeline_type,
            icon=source_pipeline.icon,
            color=source_pipeline.color,
            settings=source_pipeline.settings.copy() if source_pipeline.settings else {},
            created_by=request.user
        )
        
        # Clone fields
        for field in source_pipeline.fields.all():
            Field.objects.create(
                pipeline=cloned_pipeline,
                name=field.name,
                field_type=field.field_type,
                field_config=field.field_config.copy() if field.field_config else {},
                is_required=field.is_required,
                is_visible_in_list=field.is_visible_in_list,
                display_order=field.display_order,
                help_text=field.help_text,
                validation_rules=field.validation_rules.copy() if field.validation_rules else {},
                created_by=request.user
            )
        
        serializer = self.get_serializer(cloned_pipeline)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    @extend_schema(
        summary="Pipeline field schema",
        description="Get the field schema definition for this pipeline"
    )
    @action(detail=True, methods=['get'])
    def field_schema(self, request, pk=None):
        """Get pipeline field schema definition"""
        pipeline = self.get_object()
        
        # Check if user has BOTH pipeline read AND field read permission
        from authentication.permissions import SyncPermissionManager as PermissionManager
        permission_manager = PermissionManager(request.user)
        pipeline_access = permission_manager.has_permission('action', 'pipelines', 'read', str(pipeline.id))
        field_access = permission_manager.has_permission('action', 'fields', 'read', str(pipeline.id))
        has_field_access = pipeline_access and field_access
        
        schema = {
            'pipeline_id': pipeline.id,
            'pipeline_name': pipeline.name,
            'fields': [] if not has_field_access else []
        }
        
        if not has_field_access:
            from rest_framework.response import Response
            from rest_framework import status
            return Response({
                'detail': 'You do not have permission to view field schema for this pipeline.',
                'pipeline_id': pipeline.id,
                'pipeline_name': pipeline.name,
                'fields': []
            }, status=status.HTTP_403_FORBIDDEN)
        
        for field in pipeline.fields.all().order_by('display_order'):
            field_schema = {
                'name': field.name,
                'slug': field.slug,
                'type': field.field_type,
                'required': field.is_required,
                'visible_in_list': field.is_visible_in_list,
                'help_text': field.help_text,
                'config': field.field_config,
                'validation_rules': field.validation_rules
            }
            schema['fields'].append(field_schema)
        
        return Response(schema)
    
    def _get_active_users_count(self, pipeline):
        """Get count of users who have interacted with this pipeline recently"""
        from django.utils import timezone
        from datetime import timedelta
        
        week_ago = timezone.now() - timedelta(days=7)
        return pipeline.records.filter(
            updated_at__gte=week_ago,
            is_deleted=False
        ).values('updated_by').distinct().count()
    
    def _get_recent_activity(self, pipeline):
        """Get recent activity for the pipeline"""
        recent_records = pipeline.records.filter(
            is_deleted=False
        ).order_by('-updated_at')[:10]
        
        activity = []
        for record in recent_records:
            activity.append({
                'record_id': record.id,
                'record_title': record.title,
                'action': 'updated',
                'user': record.updated_by.email if record.updated_by else None,
                'timestamp': record.updated_at
            })
        
        return activity
    
    def _get_status_distribution(self, pipeline):
        """Get distribution of record statuses"""
        from django.db.models import Count
        
        distribution = pipeline.records.filter(
            is_deleted=False
        ).values('status').annotate(
            count=Count('id')
        ).order_by('status')
        
        return {item['status']: item['count'] for item in distribution}
    
    def _get_creation_trends(self, pipeline):
        """Get record creation trends over the last 30 days"""
        from django.utils import timezone
        from datetime import timedelta
        from django.db.models import Count
        from django.db.models.functions import TruncDate
        
        thirty_days_ago = timezone.now() - timedelta(days=30)
        
        trends = pipeline.records.filter(
            created_at__gte=thirty_days_ago,
            is_deleted=False
        ).extra({
            'date': 'date(created_at)'
        }).values('date').annotate(
            count=Count('id')
        ).order_by('date')
        
        return [
            {
                'date': item['date'],
                'count': item['count']
            }
            for item in trends
        ]
    
    def _export_csv(self, pipeline, records):
        """Export records as CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{pipeline.slug}_export.csv"'
        
        writer = csv.writer(response)
        
        # Header row
        headers = ['ID', 'Title', 'Status', 'Created At', 'Updated At']
        if has_field_access:
            for field in pipeline.fields.all().order_by('display_order'):
                headers.append(field.name)
        writer.writerow(headers)
        
        # Data rows
        for record in records:
            row = [
                record.id,
                record.title,
                record.status,
                record.created_at.isoformat(),
                record.updated_at.isoformat()
            ]
            
            if has_field_access:
                for field in pipeline.fields.all().order_by('display_order'):
                    value = record.data.get(field.slug, '')
                    if isinstance(value, (dict, list)):
                        value = json.dumps(value)
                    row.append(value)
            
            writer.writerow(row)
        
        return response
    
    def _export_json(self, pipeline, records):
        """Export records as JSON"""
        data = {
            'pipeline': {
                'id': pipeline.id,
                'name': pipeline.name,
                'exported_at': timezone.now().isoformat()
            },
            'records': []
        }
        
        for record in records:
            record_data = {
                'id': record.id,
                'title': record.title,
                'status': record.status,
                'data': record.data,
                'created_at': record.created_at.isoformat(),
                'updated_at': record.updated_at.isoformat()
            }
            data['records'].append(record_data)
        
        response = HttpResponse(
            json.dumps(data, indent=2),
            content_type='application/json'
        )
        response['Content-Disposition'] = f'attachment; filename="{pipeline.slug}_export.json"'
        
        return response
    
    def _export_xlsx(self, pipeline, records):
        """Export records as Excel file"""
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            return Response(
                {'error': 'openpyxl not installed'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = pipeline.name[:31]  # Excel sheet name limit
        
        # Headers
        headers = ['ID', 'Title', 'Status', 'Created At', 'Updated At']
        if has_field_access:
            for field in pipeline.fields.all().order_by('display_order'):
                headers.append(field.name)
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data
        for row_num, record in enumerate(records, 2):
            ws.cell(row=row_num, column=1, value=record.id)
            ws.cell(row=row_num, column=2, value=record.title)
            ws.cell(row=row_num, column=3, value=record.status)
            ws.cell(row=row_num, column=4, value=record.created_at)
            ws.cell(row=row_num, column=5, value=record.updated_at)
            
            col = 6
            if has_field_access:
                for field in pipeline.fields.all().order_by('display_order'):
                    value = record.data.get(field.slug, '')
                    if isinstance(value, (dict, list)):
                        value = json.dumps(value)
                    ws.cell(row=row_num, column=col, value=value)
                    col += 1
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{pipeline.slug}_export.xlsx"'
        
        return response


class FieldViewSet(viewsets.ModelViewSet):
    """
    Pipeline field management API
    """
    serializer_class = FieldSerializer
    permission_classes = [PipelinePermission]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['field_type', 'is_visible_in_list', 'is_searchable']
    ordering = ['display_order', 'name']
    
    def get_queryset(self):
        """Get fields for a specific pipeline"""
        pipeline_pk = self.kwargs.get('pipeline_pk')
        if pipeline_pk:
            return Field.objects.filter(
                pipeline_id=pipeline_pk
            ).select_related('pipeline', 'created_by')
        return Field.objects.none()
    
    def get_object(self):
        """Get field object, including soft-deleted fields for restore operations"""
        action = getattr(self, 'action', None)
        pipeline_pk = self.kwargs.get('pipeline_pk')
        field_pk = self.kwargs.get('pk')
        
        if not pipeline_pk or not field_pk:
            raise NotFound("Pipeline ID and Field ID are required")
        
        # For restore operations, include soft-deleted fields
        if action == 'restore':
            queryset = Field.objects.with_deleted().filter(
                pipeline_id=pipeline_pk
            ).select_related('pipeline', 'created_by')
        else:
            queryset = self.get_queryset()
        
        try:
            obj = queryset.get(pk=field_pk)
        except Field.DoesNotExist:
            raise NotFound(f"Field with ID {field_pk} not found in pipeline {pipeline_pk}")
        
        # Check object permissions
        self.check_object_permissions(self.request, obj)
        return obj
    
    def perform_create(self, serializer):
        """Create field using FieldOperationManager for migration consistency"""
        pipeline_pk = self.kwargs.get('pipeline_pk')
        pipeline = Pipeline.objects.get(id=pipeline_pk)
        
        # Use FieldOperationManager for unified field creation with migration
        from pipelines.field_operations import get_field_operation_manager
        
        field_manager = get_field_operation_manager(pipeline)
        
        # Extract field configuration from validated data
        field_config = serializer.validated_data
        field_config['pipeline'] = pipeline
        field_config['created_by'] = self.request.user
        
        # Use FieldOperationManager to create field with migration
        result = field_manager.create_field(field_config, self.request.user)
        
        if not result.success:
            from rest_framework.exceptions import ValidationError
            raise ValidationError(result.errors)
        
        # Return the created field instance for serializer
        serializer.instance = result.field
    
    def _check_business_rules_permission(self, data):
        """Check if user has permission to modify business rules"""
        if 'business_rules' in data:
            permission_manager = PermissionManager(self.request.user)
            if not permission_manager.has_permission('action', 'business_rules', 'update'):
                raise PermissionDenied(
                    "You don't have permission to modify business rules. "
                    "Required permission: business_rules.update"
                )
    
    def perform_update(self, serializer):
        """Update field using FieldOperationManager for migration consistency"""
        instance = self.get_object()
        
        # Check if business rules are being modified
        self._check_business_rules_permission(self.request.data)
        
        # Use FieldOperationManager for unified field updates with migration
        from pipelines.field_operations import get_field_operation_manager
        
        field_manager = get_field_operation_manager(instance.pipeline)
        result = field_manager.update_field(instance.id, self.request.data, self.request.user)
        
        if not result.success:
            from rest_framework.exceptions import ValidationError
            raise ValidationError(result.errors)
        
        # Update serializer instance with the updated field
        serializer.instance = result.field
    
    def perform_partial_update(self, serializer):
        """Partial update field using FieldOperationManager"""
        # FieldOperationManager handles both full and partial updates
        return self.perform_update(serializer)
    
    @extend_schema(
        summary="Reorder fields",
        description="Update the display order of multiple fields",
        request={
            "type": "object",
            "properties": {
                "field_orders": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "id": {"type": "integer"},
                            "display_order": {"type": "integer"}
                        }
                    }
                }
            }
        }
    )
    @action(detail=False, methods=['post'])
    def reorder(self, request, pipeline_pk=None):
        """Reorder pipeline fields"""
        field_orders = request.data.get('field_orders', [])
        
        updated_count = 0
        for item in field_orders:
            field_id = item.get('id')
            new_order = item.get('display_order')
            
            if field_id and new_order is not None:
                Field.objects.filter(
                    id=field_id,
                    pipeline_id=pipeline_pk
                ).update(display_order=new_order)
                updated_count += 1
        
        return Response({
            'updated_count': updated_count,
            'message': f'Updated display order for {updated_count} fields'
        })
    
    @extend_schema(
        summary="Field management actions",
        description="Perform field lifecycle management actions (soft delete, restore, schedule hard delete, impact analysis)",
        request={
            "type": "object",
            "properties": {
                "action": {
                    "type": "string",
                    "enum": ["soft_delete", "restore", "schedule_hard_delete", "impact_analysis"],
                    "description": "Action to perform on the field"
                },
                "reason": {
                    "type": "string",
                    "description": "Reason for the action (required for delete actions)"
                },
                "grace_days": {
                    "type": "integer",
                    "default": 7,
                    "description": "Days before hard delete (for schedule_hard_delete action)"
                }
            },
            "required": ["action"]
        }
    )
    @action(detail=True, methods=['post'])
    def manage(self, request, pk=None, pipeline_pk=None):
        """Perform field management actions (soft delete, restore, schedule hard delete)"""
        field = self.get_object()
        serializer = FieldManagementActionSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        action_type = serializer.validated_data['action']
        reason = serializer.validated_data.get('reason', '')
        grace_days = serializer.validated_data.get('grace_days', 7)
        
        try:
            if action_type == 'soft_delete':
                if field.is_deleted:
                    return Response({
                        'error': 'Field is already deleted'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                success, message = field.soft_delete(request.user, reason)
                if success:
                    return Response({
                        'success': True,
                        'message': message,
                        'field_status': 'soft_deleted'
                    })
                else:
                    return Response({
                        'error': message
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            elif action_type == 'restore':
                if not field.is_deleted:
                    return Response({
                        'error': 'Field is not deleted'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                success, message = field.restore(request.user)
                if success:
                    return Response({
                        'success': True,
                        'message': message,
                        'field_status': 'active'
                    })
                else:
                    return Response({
                        'error': message
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            elif action_type == 'schedule_hard_delete':
                from datetime import timedelta
                from django.utils import timezone
                delete_date = timezone.now() + timedelta(days=grace_days)
                
                success, message = field.schedule_hard_delete(request.user, reason, delete_date)
                if success:
                    return Response({
                        'success': True,
                        'message': message,
                        'field_status': 'scheduled_for_hard_delete',
                        'scheduled_date': delete_date.isoformat()
                    })
                else:
                    return Response({
                        'error': message
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            elif action_type == 'impact_analysis':
                impact = field.get_impact_analysis()
                return Response({
                    'success': True,
                    'impact_analysis': impact
                })
            
        except Exception as e:
            return Response({
                'error': f'Action failed: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @extend_schema(
        summary="Validate field migration feasibility",
        description="Check if a field migration is allowed, risky, or denied before attempting the actual migration",
        request={
            "type": "object",
            "properties": {
                "new_config": {
                    "type": "object",
                    "description": "Proposed new field configuration"
                },
                "include_impact_preview": {
                    "type": "boolean",
                    "default": False,
                    "description": "Include detailed impact analysis in response"
                }
            },
            "required": ["new_config"]
        },
        responses={
            200: {
                "type": "object",
                "properties": {
                    "validation": {
                        "type": "object",
                        "properties": {
                            "allowed": {"type": "boolean"},
                            "category": {"type": "string", "enum": ["safe", "risky", "denied"]},
                            "risk_level": {"type": "string", "enum": ["low", "medium", "high"]},
                            "reason": {"type": "string"},
                            "alternatives": {"type": "array", "items": {"type": "string"}}
                        }
                    }
                }
            }
        }
    )
    @action(detail=True, methods=['post'])
    def validate_migration(self, request, pk=None, pipeline_pk=None):
        """Validate field migration feasibility with comprehensive analysis"""
        field = self.get_object()
        
        # Debug logging
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"🔄 Migration validation request data: {request.data}")
        
        # Import here to avoid circular imports
        from pipelines.serializers import MigrationValidationSerializer
        from pipelines.validation import MigrationValidator
        
        serializer = MigrationValidationSerializer(data=request.data)
        
        if not serializer.is_valid():
            logger.error(f"❌ Serializer validation failed: {serializer.errors}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        new_config = serializer.validated_data['new_config']
        include_impact = serializer.validated_data.get('include_impact_preview', False)
        
        try:
            # Core validation
            validation = MigrationValidator.validate_field_change(field, new_config)
            
            # Enhanced response with comprehensive analysis
            response_data = {
                'validation': validation,
                'field_info': {
                    'id': field.id,
                    'name': field.name,
                    'display_name': field.display_name or field.name,
                    'current_type': field.field_type,
                    'target_type': new_config.get('field_type', field.field_type),
                    'pipeline_id': field.pipeline.id,
                    'pipeline_name': field.pipeline.name
                }
            }
            
            # Add comprehensive analysis for allowed migrations
            if validation.get('allowed'):
                # Performance estimation
                performance = MigrationValidator.estimate_performance(field, new_config)
                response_data['performance_estimate'] = performance
                
                # Data transformation preview
                data_preview = MigrationValidator.generate_data_preview(field, new_config)
                response_data['data_preview'] = data_preview
                
                # Dependency analysis
                dependencies = MigrationValidator.analyze_dependencies(field)
                response_data['dependency_analysis'] = dependencies
                
                # Include migration impact if requested
                if include_impact:
                    from pipelines.migrator import FieldSchemaMigrator
                    migrator = FieldSchemaMigrator(field.pipeline)
                    impact = migrator.analyze_field_change_impact(field, new_config)
                    response_data['migration_impact'] = impact
                
                # Configuration requirements for target field type
                target_type = new_config.get('field_type', field.field_type)
                response_data['configuration_requirements'] = {
                    'target_field_type': target_type,
                    'required_config_keys': self._get_required_config_keys(target_type),
                    'optional_config_keys': self._get_optional_config_keys(target_type),
                    'validation_rules': self._get_validation_rules_for_type(target_type)
                }
            
            # Add alternatives for denied migrations
            elif not validation.get('allowed'):
                alternatives = MigrationValidator.get_migration_alternatives(
                    field.field_type, 
                    new_config.get('field_type', field.field_type)
                )
                response_data['alternatives'] = alternatives
            
            return Response(response_data)
            
        except Exception as e:
            logger.error(f"Migration validation error: {e}")
            return Response({
                'error': f'Validation failed: {str(e)}',
                'field_id': field.id
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def _get_required_config_keys(self, field_type):
        """Get required configuration keys for a field type"""
        config_requirements = {
            'select': ['options'],
            'multiselect': ['options'],
            'ai_generated': ['prompt', 'model'],
            'computed': ['formula'],
            'relation': ['target_pipeline'],
            'number': [],
            'text': [],
            'textarea': [],
            'email': [],
            'url': [],
            'phone': [],
            'date': [],
            'datetime': [],
            'boolean': [],
            'file': [],
            'decimal': []
        }
        return config_requirements.get(field_type, [])
    
    def _get_optional_config_keys(self, field_type):
        """Get optional configuration keys for a field type"""
        optional_config = {
            'select': ['allow_other', 'default_value'],
            'multiselect': ['allow_other', 'default_values', 'max_selections'],
            'ai_generated': ['temperature', 'max_tokens', 'tools_enabled', 'cache_duration'],
            'computed': ['format', 'update_trigger'],
            'relation': ['display_field', 'allow_multiple'],
            'number': ['min_value', 'max_value', 'default_value'],
            'text': ['max_length', 'default_value', 'placeholder'],
            'textarea': ['max_length', 'default_value', 'placeholder', 'rows'],
            'email': ['default_value', 'placeholder'],
            'url': ['default_value', 'placeholder'],
            'phone': ['default_value', 'placeholder', 'format'],
            'date': ['default_value', 'min_date', 'max_date'],
            'datetime': ['default_value', 'min_datetime', 'max_datetime'],
            'boolean': ['default_value'],
            'file': ['allowed_types', 'max_size'],
            'decimal': ['min_value', 'max_value', 'decimal_places', 'default_value']
        }
        return optional_config.get(field_type, [])
    
    def _get_validation_rules_for_type(self, field_type):
        """Get validation rules that apply to a field type"""
        validation_rules = {
            'text': ['required', 'min_length', 'max_length', 'pattern'],
            'textarea': ['required', 'min_length', 'max_length'],
            'number': ['required', 'min_value', 'max_value'],
            'decimal': ['required', 'min_value', 'max_value', 'decimal_places'],
            'email': ['required', 'pattern'],
            'url': ['required', 'pattern'],
            'phone': ['required', 'pattern'],
            'date': ['required', 'min_date', 'max_date'],
            'datetime': ['required', 'min_datetime', 'max_datetime'],
            'boolean': ['required'],
            'select': ['required'],
            'multiselect': ['required', 'min_selections', 'max_selections'],
            'file': ['required', 'max_size', 'allowed_types'],
            'ai_generated': ['required'],
            'computed': [],
            'relation': ['required']
        }
        return validation_rules.get(field_type, ['required'])
    
    @extend_schema(
        summary="List deleted fields",
        description="Get all soft-deleted fields for the pipeline that can be restored",
        responses={
            200: {
                "type": "array",
                "items": {"$ref": "#/components/schemas/Field"}
            }
        }
    )
    @action(detail=False, methods=['get'])
    def deleted(self, request, pipeline_pk=None):
        """Get all soft-deleted fields for this pipeline"""
        deleted_fields = Field.objects.with_deleted().filter(
            pipeline_id=pipeline_pk,
            is_deleted=True
        ).order_by('-deleted_at')
        
        serializer = self.get_serializer(deleted_fields, many=True)
        return Response(serializer.data)
    
    @extend_schema(
        summary="Restore field",
        description="Restore a soft-deleted field with validation and dry-run support",
        request={
            "type": "object",
            "properties": {
                "reason": {
                    "type": "string",
                    "description": "Reason for restoring the field"
                },
                "dry_run": {
                    "type": "boolean",
                    "default": False,
                    "description": "Preview restore impact without performing actual restore"
                },
                "force": {
                    "type": "boolean",
                    "default": False,
                    "description": "Force restore even if validation warnings exist"
                }
            }
        }
    )
    @action(detail=True, methods=['post'])
    def restore(self, request, pk=None, pipeline_pk=None):
        """Enhanced restore with validation and dry-run support"""
        field = self.get_object()
        
        # Import serializer here to avoid circular imports
        from pipelines.serializers import FieldRestoreSerializer
        serializer = FieldRestoreSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        reason = serializer.validated_data.get('reason', '')
        dry_run = serializer.validated_data.get('dry_run', False)
        force = serializer.validated_data.get('force', False)
        
        try:
            result = field.restore_with_validation(
                user=request.user,
                force=force,
                dry_run=dry_run
            )
            
            if result['success']:
                status_code = status.HTTP_200_OK
            else:
                status_code = status.HTTP_400_BAD_REQUEST if not dry_run else status.HTTP_200_OK
            
            return Response(result, status=status_code)
            
        except Exception as e:
            return Response({
                'success': False,
                'errors': [f'Restore failed: {str(e)}']
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @extend_schema(
        summary="Bulk restore fields",
        description="Restore multiple soft-deleted fields in batch",
        request={
            "type": "object",
            "properties": {
                "field_ids": {
                    "type": "array",
                    "items": {"type": "integer"},
                    "description": "List of field IDs to restore (max 20)"
                },
                "reason": {
                    "type": "string",
                    "description": "Reason for bulk restore"
                },
                "force": {
                    "type": "boolean",
                    "default": False,
                    "description": "Force restore for all fields even with validation warnings"
                }
            },
            "required": ["field_ids"]
        }
    )
    @action(detail=False, methods=['post'])
    def bulk_restore(self, request, pipeline_pk=None):
        """Bulk restore multiple fields"""
        from pipelines.serializers import BulkFieldRestoreSerializer
        serializer = BulkFieldRestoreSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        field_ids = serializer.validated_data['field_ids']
        reason = serializer.validated_data.get('reason', '')
        force = serializer.validated_data.get('force', False)
        
        # Get fields (they're already validated by serializer)
        fields = Field.objects.with_deleted().filter(
            id__in=field_ids,
            pipeline_id=pipeline_pk
        )
        
        results = []
        successful_count = 0
        failed_count = 0
        
        for field in fields:
            try:
                result = field.restore_with_validation(
                    user=request.user,
                    force=force,
                    dry_run=False
                )
                
                result['field_id'] = field.id
                result['field_name'] = field.name
                results.append(result)
                
                if result['success']:
                    successful_count += 1
                else:
                    failed_count += 1
                    
            except Exception as e:
                results.append({
                    'field_id': field.id,
                    'field_name': field.name,
                    'success': False,
                    'errors': [f'Restore failed: {str(e)}']
                })
                failed_count += 1
        
        return Response({
            'success': failed_count == 0,
            'total_fields': len(field_ids),
            'successful_count': successful_count,
            'failed_count': failed_count,
            'results': results
        })
    
    @extend_schema(
        summary="Migrate field schema",
        description="Perform field schema migration with data transformation. Validates migration feasibility first.",
        request={
            "type": "object",
            "properties": {
                "new_config": {
                    "type": "object",
                    "description": "New field configuration to migrate to"
                },
                "dry_run": {
                    "type": "boolean",
                    "default": False,
                    "description": "Perform dry run without making actual changes"
                },
                "batch_size": {
                    "type": "integer",
                    "default": 100,
                    "minimum": 10,
                    "maximum": 1000,
                    "description": "Number of records to process per batch"
                },
                "force": {
                    "type": "boolean",
                    "default": False,
                    "description": "Force migration even for risky changes (requires explicit confirmation)"
                }
            },
            "required": ["new_config"]
        }
    )
    @action(detail=True, methods=['post'])
    def migrate_schema(self, request, pk=None, pipeline_pk=None):
        """Migrate field schema with comprehensive validation and safety checks"""
        field = self.get_object()
        
        # Import here to avoid circular imports
        from pipelines.serializers import FieldMigrationSerializer
        from pipelines.validation import MigrationValidator
        from pipelines.migrator import FieldSchemaMigrator
        
        serializer = FieldMigrationSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        new_config = serializer.validated_data['new_config']
        dry_run = serializer.validated_data.get('dry_run', False)
        batch_size = serializer.validated_data.get('batch_size', 100)
        force = request.data.get('force', False)
        
        try:
            # First, validate migration feasibility
            validation = MigrationValidator.validate_field_change(field, new_config)
            
            if not validation['allowed']:
                return Response({
                    'error': 'Migration denied',
                    'validation': validation,
                    'message': f"Migration blocked: {validation['reason']}"
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Check if risky migration requires force flag
            if validation['category'] == 'risky' and not force:
                return Response({
                    'error': 'Risky migration requires confirmation',
                    'validation': validation,
                    'message': 'Add "force": true to proceed with this risky migration',
                    'required_confirmation': True
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Initialize migrator and analyze impact
            migrator = FieldSchemaMigrator(field.pipeline)
            impact = migrator.analyze_field_change_impact(field, new_config)
            
            # Handle dry run
            if dry_run:
                return Response({
                    'success': True,
                    'dry_run': True,
                    'validation': validation,
                    'impact_analysis': impact,
                    'message': 'Dry run completed - no changes made'
                })
            
            # Perform the migration
            if impact.get('migration_required', False):
                # For actual migrations with data changes, use background task
                from pipelines.tasks import migrate_field_schema
                task = migrate_field_schema.delay(
                    field.pipeline.id,
                    field.slug,
                    new_config,
                    batch_size
                )
                
                return Response({
                    'success': True,
                    'migration_started': True,
                    'task_id': task.id,
                    'validation': validation,
                    'impact_analysis': impact,
                    'message': 'Migration started in background',
                    'status_check_url': f'/api/pipelines/{pipeline_pk}/fields/{pk}/migration_status/?task_id={task.id}'
                })
            else:
                # Simple configuration update - no data migration needed
                # Update field configuration directly
                if 'field_type' in new_config:
                    field.field_type = new_config['field_type']
                if 'field_config' in new_config:
                    field.field_config = new_config['field_config']
                if 'storage_constraints' in new_config:
                    field.storage_constraints = new_config['storage_constraints']
                if 'business_rules' in new_config:
                    field.business_rules = new_config['business_rules']
                    
                field.save()
                
                return Response({
                    'success': True,
                    'migration_completed': True,
                    'validation': validation,
                    'impact_analysis': impact,
                    'message': 'Configuration updated successfully - no data migration required'
                })
                
        except Exception as e:
            return Response({
                'error': f'Migration failed: {str(e)}',
                'field_id': field.id
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)