"""
Views for pipeline system API
"""
from rest_framework import viewsets, permissions, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from django.shortcuts import get_object_or_404
from django.db.models import Q
from django_filters.rest_framework import DjangoFilterBackend
import json
import csv
import io
from openpyxl import Workbook
from django.http import HttpResponse

from .models import Pipeline, Field, Record, PipelineTemplate
from .serializers import (
    PipelineSerializer, PipelineCreateSerializer, FieldSerializer,
    RecordSerializer, RecordCreateSerializer, PipelineTemplateSerializer,
    PipelineTemplateCreatePipelineSerializer, FieldValidationSerializer,
    RecordSearchSerializer, BulkRecordActionSerializer,
    FieldManagementActionSerializer, FieldMigrationSerializer
)
# Legacy AI processor removed - using new ai/integrations.py system
from api.permissions import PipelinePermission, RecordPermission
from authentication.permissions import SyncPermissionManager


class PipelineViewSet(viewsets.ModelViewSet):
    """ViewSet for managing pipelines"""
    queryset = Pipeline.objects.all()
    permission_classes = [PipelinePermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['pipeline_type', 'is_active', 'access_level']
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created_at', 'updated_at', 'record_count']
    ordering = ['-updated_at']
    
    def get_serializer_class(self):
        if self.action == 'create':
            return PipelineCreateSerializer
        return PipelineSerializer
    
    def get_queryset(self):
        """Filter pipelines based on user permissions"""
        user = self.request.user
        if user.is_superuser:
            return Pipeline.objects.all()
        
        # Use comprehensive permission system
        permission_manager = SyncPermissionManager(user)
        
        # Check if user has read_all permission for pipelines
        if permission_manager.has_permission('action', 'pipelines', 'read_all'):
            return Pipeline.objects.all()
        
        # Filter to only pipelines user has access to
        accessible_pipeline_ids = self._get_accessible_pipeline_ids(user)
        return Pipeline.objects.filter(id__in=accessible_pipeline_ids)
    
    def _get_accessible_pipeline_ids(self, user):
        """Get pipeline IDs that user has access to"""
        permission_manager = SyncPermissionManager(user)
        accessible_ids = []
        
        # Check each pipeline for read access
        for pipeline in Pipeline.objects.all():
            if (permission_manager.has_permission('action', 'pipelines', 'read', str(pipeline.id)) or
                permission_manager.has_permission('action', 'pipelines', 'read') or
                pipeline.created_by == user or
                pipeline.access_level in ['public', 'internal']):
                accessible_ids.append(pipeline.id)
        
        return accessible_ids
    
    @action(detail=True, methods=['post'])
    def clone(self, request, pk=None):
        """Clone a pipeline"""
        pipeline = self.get_object()
        name = request.data.get('name')
        
        if not name:
            return Response(
                {'error': 'Name is required for cloning'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            cloned_pipeline = pipeline.clone(name, request.user)
            serializer = PipelineSerializer(cloned_pipeline)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['get'])
    def fields(self, request, pk=None):
        """Get pipeline fields"""
        pipeline = self.get_object()
        fields = pipeline.fields.all()
        serializer = FieldSerializer(fields, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def add_field(self, request, pk=None):
        """Add a field to the pipeline"""
        pipeline = self.get_object()
        serializer = FieldSerializer(data=request.data)
        
        if serializer.is_valid():
            serializer.save(pipeline=pipeline, created_by=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['get'])
    def records(self, request, pk=None):
        """Get pipeline records"""
        pipeline = self.get_object()
        records = pipeline.records.filter(is_deleted=False)
        
        # Apply search filters
        search_serializer = RecordSearchSerializer(
            data=request.query_params,
            pipeline=pipeline
        )
        
        if search_serializer.is_valid():
            data = search_serializer.validated_data
            
            # Apply filters
            if data.get('q'):
                records = records.extra(
                    where=["search_vector @@ plainto_tsquery(%s)"],
                    params=[data['q']]
                )
            
            if data.get('status'):
                records = records.filter(status=data['status'])
            
            if data.get('tags'):
                records = records.filter(tags__overlap=data['tags'])
            
            if data.get('created_after'):
                records = records.filter(created_at__gte=data['created_after'])
            
            if data.get('created_before'):
                records = records.filter(created_at__lte=data['created_before'])
            
            # Apply dynamic field filters
            for field in pipeline.fields.filter(is_searchable=True):
                field_name = f"field_{field.slug}"
                
                if field_name in data:
                    records = records.filter(
                        **{f"data__{field.slug}__icontains": data[field_name]}
                    )
        
        # Paginate results
        page = self.paginate_queryset(records)
        if page is not None:
            serializer = RecordSerializer(page, many=True, context={'request': request})
            return self.get_paginated_response(serializer.data)
        
        serializer = RecordSerializer(records, many=True, context={'request': request})
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def create_record(self, request, pk=None):
        """Create a new record in the pipeline"""
        pipeline = self.get_object()
        serializer = RecordCreateSerializer(
            data=request.data,
            context={'request': request, 'pipeline': pipeline}
        )
        
        if serializer.is_valid():
            record = serializer.save()
            
            # Process AI fields asynchronously if present
            # COMMENTED OUT: Legacy AI processing - now handled automatically via Record.save()
            # if pipeline.get_ai_fields().exists():
            #     try:
            #         ai_results = process_ai_fields_sync(record)
            #         if ai_results:
            #             record.data.update(ai_results)
            #             record.save(update_fields=['data'])
            #     except Exception as e:
            #         # AI processing failed, but record was created successfully
            #         pass
            
            response_serializer = RecordSerializer(record, context={'request': request})
            return Response(response_serializer.data, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['get'])
    def export(self, request, pk=None):
        """Export pipeline records"""
        pipeline = self.get_object()
        export_format = request.query_params.get('format', 'csv')
        
        records = pipeline.records.filter(is_deleted=False)
        
        if export_format == 'json':
            data = []
            for record in records:
                record_data = record.to_dict(include_metadata=True)
                data.append(record_data)
            
            response = HttpResponse(
                json.dumps(data, indent=2),
                content_type='application/json'
            )
            response['Content-Disposition'] = f'attachment; filename="{pipeline.slug}_records.json"'
            
        elif export_format == 'xlsx':
            wb = Workbook()
            ws = wb.active
            ws.title = pipeline.name
            
            # Write headers
            headers = ['ID', 'Title', 'Status', 'Created At', 'Updated At']
            field_headers = [field.name for field in pipeline.fields.all()]
            headers.extend(field_headers)
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Write data
            for row, record in enumerate(records, 2):
                ws.cell(row=row, column=1, value=record.id)
                ws.cell(row=row, column=2, value=record.title)
                ws.cell(row=row, column=3, value=record.status)
                ws.cell(row=row, column=4, value=record.created_at)
                ws.cell(row=row, column=5, value=record.updated_at)
                
                for col, field in enumerate(pipeline.fields.all(), 6):
                    value = record.data.get(field.slug, '')
                    ws.cell(row=row, column=col, value=str(value))
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{pipeline.slug}_records.xlsx"'
            
        else:  # CSV format
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write headers
            headers = ['ID', 'Title', 'Status', 'Created At', 'Updated At']
            field_headers = [field.name for field in pipeline.fields.all()]
            headers.extend(field_headers)
            writer.writerow(headers)
            
            # Write data
            for record in records:
                row = [
                    record.id,
                    record.title,
                    record.status,
                    record.created_at,
                    record.updated_at
                ]
                
                for field in pipeline.fields.all():
                    value = record.data.get(field.slug, '')
                    row.append(str(value))
                
                writer.writerow(row)
            
            response = HttpResponse(output.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{pipeline.slug}_records.csv"'
        
        return response


class FieldViewSet(viewsets.ModelViewSet):
    """Enhanced ViewSet for managing pipeline fields with lifecycle management"""
    serializer_class = FieldSerializer
    permission_classes = [PipelinePermission]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['pipeline', 'field_type', 'is_ai_field', 'is_deleted']
    ordering_fields = ['display_order', 'name', 'created_at']
    ordering = ['display_order']
    
    def get_queryset(self):
        """Filter fields based on pipeline access and include deleted fields for admin"""
        user = self.request.user
        
        # Check if we should include deleted fields
        include_deleted = self.request.query_params.get('include_deleted', 'false').lower() == 'true'
        
        if user.is_superuser:
            base_queryset = Field.objects.with_deleted() if include_deleted else Field.objects.all()
        else:
            # Use comprehensive permission system
            permission_manager = SyncPermissionManager(user)
            
            # Get accessible pipelines and filter fields accordingly
            if permission_manager.has_permission('action', 'pipelines', 'read_all'):
                accessible_pipelines = Pipeline.objects.all()
            else:
                accessible_pipeline_ids = self._get_accessible_pipeline_ids(user, permission_manager)
                accessible_pipelines = Pipeline.objects.filter(id__in=accessible_pipeline_ids)
            
            if include_deleted:
                base_queryset = Field.objects.with_deleted().filter(pipeline__in=accessible_pipelines)
            else:
                base_queryset = Field.objects.filter(pipeline__in=accessible_pipelines)
        
        return base_queryset
    
    def _get_accessible_pipeline_ids(self, user, permission_manager):
        """Get pipeline IDs that user has access to"""
        accessible_ids = []
        
        for pipeline in Pipeline.objects.all():
            if (permission_manager.has_permission('action', 'pipelines', 'read', str(pipeline.id)) or
                permission_manager.has_permission('action', 'pipelines', 'read') or
                pipeline.created_by == user or
                pipeline.access_level in ['public', 'internal']):
                accessible_ids.append(pipeline.id)
        
        return accessible_ids
    
    def create(self, request, *args, **kwargs):
        """Create a new field using FieldOperationManager"""
        # Extract pipeline from request data
        pipeline_id = request.data.get('pipeline')
        if not pipeline_id:
            return Response({
                'success': False,
                'errors': ['Pipeline ID is required']
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            pipeline = Pipeline.objects.get(id=pipeline_id)
        except Pipeline.DoesNotExist:
            return Response({
                'success': False,
                'errors': ['Pipeline not found']
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Use FieldOperationManager for unified field creation
        try:
            from .field_operations import get_field_operation_manager
            
            field_manager = get_field_operation_manager(pipeline)
            result = field_manager.create_field(request.data, request.user)
            
            if result.success:
                # Return success response with created field
                serializer = self.get_serializer(result.field)
                return Response({
                    'success': True,
                    'field': serializer.data,
                    'operation_id': result.operation_id,
                    'warnings': result.warnings,
                    'metadata': result.metadata
                }, status=status.HTTP_201_CREATED)
            else:
                # Return validation errors
                return Response({
                    'success': False,
                    'errors': result.errors,
                    'operation_id': result.operation_id,
                    'warnings': result.warnings
                }, status=status.HTTP_400_BAD_REQUEST)
                
        except Exception as e:
            return Response({
                'success': False,
                'errors': [f'Field creation failed: {str(e)}']
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def update(self, request, *args, **kwargs):
        """Update field using FieldOperationManager"""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        
        # Use FieldOperationManager for unified field updates
        try:
            from .field_operations import get_field_operation_manager
            
            field_manager = get_field_operation_manager(instance.pipeline)
            result = field_manager.update_field(instance.id, request.data, request.user)
            
            if result.success:
                # Return success response with updated field
                serializer = self.get_serializer(result.field)
                return Response({
                    'success': True,
                    'field': serializer.data,
                    'operation_id': result.operation_id,
                    'warnings': result.warnings,
                    'metadata': result.metadata
                })
            else:
                # Return validation/migration errors
                return Response({
                    'success': False,
                    'errors': result.errors,
                    'operation_id': result.operation_id,
                    'warnings': result.warnings
                }, status=status.HTTP_400_BAD_REQUEST)
                
        except Exception as e:
            return Response({
                'success': False,
                'errors': [f'Field update failed: {str(e)}']
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def destroy(self, request, *args, **kwargs):
        """Soft delete field using FieldOperationManager"""
        instance = self.get_object()
        
        # Check if already deleted
        if instance.is_deleted:
            return Response({
                'success': False,
                'error': 'Field is already deleted'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Use FieldOperationManager for unified field deletion
        try:
            from .field_operations import get_field_operation_manager
            
            field_manager = get_field_operation_manager(instance.pipeline)
            result = field_manager.delete_field(instance.id, request.user, hard_delete=False)
            
            if result.success:
                return Response({
                    'success': True,
                    'message': 'Field soft deleted successfully',
                    'operation_id': result.operation_id,
                    'warnings': result.warnings,
                    'metadata': result.metadata
                }, status=status.HTTP_204_NO_CONTENT)
            else:
                return Response({
                    'success': False,
                    'errors': result.errors,
                    'operation_id': result.operation_id
                }, status=status.HTTP_400_BAD_REQUEST)
                
        except Exception as e:
            return Response({
                'success': False,
                'errors': [f'Field deletion failed: {str(e)}']
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['post'])
    def validate_value(self, request, pk=None):
        """Validate a value against the field"""
        field = self.get_object()
        serializer = FieldValidationSerializer(
            data=request.data,
            context={'field': field}
        )
        
        if serializer.is_valid():
            return Response({
                'is_valid': True,
                'cleaned_value': serializer.validated_data['cleaned_value']
            })
        
        return Response({
            'is_valid': False,
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def manage(self, request, pk=None):
        """Perform field management actions (soft delete, restore, schedule hard delete)"""
        field = self.get_object()
        serializer = FieldManagementActionSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        action_type = serializer.validated_data['action']
        reason = serializer.validated_data.get('reason', '')
        grace_days = serializer.validated_data.get('grace_days', 7)
        
        try:
            if action_type == 'soft_delete':
                if field.is_deleted:
                    return Response({
                        'error': 'Field is already deleted'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                success, message = field.soft_delete(request.user, reason)
                if success:
                    return Response({
                        'success': True,
                        'message': message,
                        'field_status': 'soft_deleted'
                    })
                else:
                    return Response({
                        'error': message
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            elif action_type == 'restore':
                if not field.is_deleted:
                    return Response({
                        'error': 'Field is not deleted'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                success, message = field.restore(request.user)
                if success:
                    return Response({
                        'success': True,
                        'message': message,
                        'field_status': 'active'
                    })
                else:
                    return Response({
                        'error': message
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            elif action_type == 'schedule_hard_delete':
                from datetime import timedelta
                from django.utils import timezone
                delete_date = timezone.now() + timedelta(days=grace_days)
                
                success, message = field.schedule_hard_delete(request.user, reason, delete_date)
                if success:
                    return Response({
                        'success': True,
                        'message': message,
                        'field_status': 'scheduled_for_hard_delete',
                        'scheduled_date': delete_date.isoformat()
                    })
                else:
                    return Response({
                        'error': message
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            elif action_type == 'impact_analysis':
                impact = field.get_impact_analysis()
                return Response({
                    'success': True,
                    'impact_analysis': impact
                })
            
        except Exception as e:
            return Response({
                'error': f'Action failed: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['post'])
    def migrate_schema(self, request, pk=None):
        """Migrate field schema changes using FieldOperationManager"""
        field = self.get_object()
        serializer = FieldMigrationSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        new_config = serializer.validated_data['new_config']
        dry_run = serializer.validated_data['dry_run']
        batch_size = serializer.validated_data.get('batch_size', 1000)
        
        try:
            from .field_operations import get_field_operation_manager
            
            field_manager = get_field_operation_manager(field.pipeline)
            
            if dry_run:
                # For dry run, use FieldValidator to analyze changes without applying them
                from .validation.field_validator import FieldValidator as AdvancedFieldValidator
                validator = AdvancedFieldValidator()
                
                # Validate the changes
                validation_result = validator.validate_field_update(field, new_config)
                
                if not validation_result.valid:
                    return Response({
                        'success': False,
                        'dry_run': True,
                        'errors': validation_result.errors,
                        'warnings': validation_result.warnings
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Analyze impact using state manager
                from .state.field_state_manager import get_field_state_manager
                state_manager = get_field_state_manager()
                
                # Generate temporary operation ID for dry run analysis
                import uuid
                temp_operation_id = f"dry_run_{uuid.uuid4().hex[:8]}"
                
                # Capture current state and analyze changes
                state_manager.capture_field_state(field.id, temp_operation_id)
                
                # Create temporary field with new config for analysis
                temp_field = field
                for key, value in new_config.items():
                    if hasattr(temp_field, key):
                        setattr(temp_field, key, value)
                
                change_analysis = state_manager.get_field_changes(field.id, temp_field, temp_operation_id)
                
                # Clean up temporary state
                state_manager.cleanup_operation_state(temp_operation_id)
                
                return Response({
                    'success': True,
                    'dry_run': True,
                    'impact_analysis': change_analysis,
                    'validation_warnings': validation_result.warnings,
                    'message': 'Dry run completed - no changes made'
                })
            
            # Perform actual migration using FieldOperationManager
            result = field_manager.update_field(field.id, new_config, request.user)
            
            if result.success:
                return Response({
                    'success': True,
                    'field': self.get_serializer(result.field).data,
                    'operation_id': result.operation_id,
                    'warnings': result.warnings,
                    'metadata': result.metadata,
                    'message': 'Field updated successfully with automatic migration handling'
                })
            else:
                return Response({
                    'success': False,
                    'errors': result.errors,
                    'operation_id': result.operation_id,
                    'warnings': result.warnings
                }, status=status.HTTP_400_BAD_REQUEST)
        
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Migration failed: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def deleted(self, request):
        """Get all soft deleted fields"""
        user = request.user
        if not user.is_superuser:
            return Response({
                'error': 'Only administrators can view deleted fields'
            }, status=status.HTTP_403_FORBIDDEN)
        
        deleted_fields = Field.objects.deleted_only()
        serializer = self.get_serializer(deleted_fields, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def scheduled_for_deletion(self, request):
        """Get all fields scheduled for hard deletion"""
        user = request.user
        if not user.is_superuser:
            return Response({
                'error': 'Only administrators can view scheduled deletions'
            }, status=status.HTTP_403_FORBIDDEN)
        
        scheduled_fields = Field.objects.scheduled_for_hard_delete()
        serializer = self.get_serializer(scheduled_fields, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def migration_status(self, request):
        """Check the status of field migration tasks"""
        task_id = request.query_params.get('task_id')
        if not task_id:
            return Response({
                'error': 'task_id parameter is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            from django.core.cache import cache
            from celery import current_app
            
            # Check if task exists
            task = current_app.AsyncResult(task_id)
            
            if task.state == 'PENDING':
                return Response({
                    'task_id': task_id,
                    'status': 'pending',
                    'message': 'Task is waiting to be processed'
                })
            elif task.state == 'PROGRESS':
                return Response({
                    'task_id': task_id,
                    'status': 'in_progress',
                    'message': 'Task is currently being processed',
                    'info': task.info
                })
            elif task.state == 'SUCCESS':
                # Try to get cached result
                cache_key = f"field_migration:{task_id}"
                result = cache.get(cache_key, task.result)
                
                return Response({
                    'task_id': task_id,
                    'status': 'completed',
                    'result': result
                })
            elif task.state == 'FAILURE':
                return Response({
                    'task_id': task_id,
                    'status': 'failed',
                    'error': str(task.info),
                    'message': 'Task failed to complete'
                })
            else:
                return Response({
                    'task_id': task_id,
                    'status': task.state.lower(),
                    'message': f'Task is in {task.state} state'
                })
        
        except Exception as e:
            return Response({
                'error': f'Failed to check task status: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'])
    def bulk_operations(self, request):
        """Perform bulk field operations using FieldOperationManager"""
        operation = request.data.get('operation')
        field_ids = request.data.get('field_ids', [])
        pipeline_id = request.data.get('pipeline_id')
        
        if not operation:
            return Response({
                'success': False,
                'errors': ['Operation is required']
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not field_ids:
            return Response({
                'success': False,
                'errors': ['Field IDs are required']
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if operation in ['restore'] and not pipeline_id:
            return Response({
                'success': False,
                'errors': ['Pipeline ID is required for bulk restore operations']
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Get fields user has access to
            if operation == 'restore':
                # For restore, include deleted fields
                fields = self.get_queryset().filter(
                    id__in=field_ids,
                    is_deleted=True
                ).select_related('pipeline')
            else:
                # For other operations, only active fields
                fields = self.get_queryset().filter(
                    id__in=field_ids,
                    is_deleted=False
                ).select_related('pipeline')
            
            if not fields.exists():
                return Response({
                    'success': False,
                    'errors': ['No accessible fields found']
                }, status=status.HTTP_404_NOT_FOUND)
            
            results = {
                'success': True,
                'operation': operation,
                'results': [],
                'summary': {'processed': 0, 'successful': 0, 'failed': 0}
            }
            
            # Group fields by pipeline for efficient operation management
            pipeline_fields = {}
            for field in fields:
                if field.pipeline.id not in pipeline_fields:
                    pipeline_fields[field.pipeline.id] = []
                pipeline_fields[field.pipeline.id].append(field)
            
            # Process each pipeline group
            for pipeline_id, field_list in pipeline_fields.items():
                try:
                    from .field_operations import get_field_operation_manager
                    field_manager = get_field_operation_manager(field_list[0].pipeline)
                    
                    for field in field_list:
                        field_result = {
                            'field_id': field.id,
                            'field_name': field.name,
                            'success': False,
                            'errors': [],
                            'warnings': [],
                            'operation_id': None
                        }
                        
                        try:
                            if operation == 'soft_delete':
                                result = field_manager.delete_field(field.id, request.user, hard_delete=False)
                            elif operation == 'restore':
                                result = field_manager.restore_field(field.id, request.user)
                            elif operation == 'hard_delete':
                                result = field_manager.delete_field(field.id, request.user, hard_delete=True)
                            else:
                                field_result['errors'] = [f'Unknown operation: {operation}']
                                results['results'].append(field_result)
                                results['summary']['failed'] += 1
                                continue
                            
                            field_result['success'] = result.success
                            field_result['operation_id'] = result.operation_id
                            field_result['errors'] = result.errors
                            field_result['warnings'] = result.warnings
                            
                            if result.success:
                                results['summary']['successful'] += 1
                            else:
                                results['summary']['failed'] += 1
                            
                            results['summary']['processed'] += 1
                            
                        except Exception as e:
                            field_result['errors'] = [f'Operation failed: {str(e)}']
                            results['summary']['failed'] += 1
                            results['summary']['processed'] += 1
                        
                        results['results'].append(field_result)
                    
                except Exception as e:
                    # Pipeline-level error
                    for field in field_list:
                        field_result = {
                            'field_id': field.id,
                            'field_name': field.name,
                            'success': False,
                            'errors': [f'Pipeline operation failed: {str(e)}'],
                            'warnings': [],
                            'operation_id': None
                        }
                        results['results'].append(field_result)
                        results['summary']['failed'] += 1
                        results['summary']['processed'] += 1
            
            return Response(results)
            
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Bulk operation failed: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class RecordViewSet(viewsets.ModelViewSet):
    """ViewSet for managing pipeline records"""
    queryset = Record.objects.filter(is_deleted=False)
    serializer_class = RecordSerializer
    permission_classes = [RecordPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['pipeline', 'status']
    search_fields = ['title', 'data']
    ordering_fields = ['created_at', 'updated_at', 'title']
    ordering = ['-updated_at']
    
    def get_queryset(self):
        """Filter records based on pipeline access"""
        user = self.request.user
        if user.is_superuser:
            return Record.objects.filter(is_deleted=False)
        
        # Use comprehensive permission system
        permission_manager = SyncPermissionManager(user)
        
        # Get accessible pipelines and filter records accordingly
        if permission_manager.has_permission('action', 'pipelines', 'read_all'):
            accessible_pipelines = Pipeline.objects.all()
        else:
            accessible_pipeline_ids = self._get_accessible_pipeline_ids(user, permission_manager)
            accessible_pipelines = Pipeline.objects.filter(id__in=accessible_pipeline_ids)
        
        return Record.objects.filter(is_deleted=False, pipeline__in=accessible_pipelines)
    
    def _get_accessible_pipeline_ids(self, user, permission_manager):
        """Get pipeline IDs that user has access to"""
        accessible_ids = []
        
        for pipeline in Pipeline.objects.all():
            if (permission_manager.has_permission('action', 'pipelines', 'read', str(pipeline.id)) or
                permission_manager.has_permission('action', 'pipelines', 'read') or
                pipeline.created_by == user or
                pipeline.access_level in ['public', 'internal']):
                accessible_ids.append(pipeline.id)
        
        return accessible_ids
    
    def get_serializer_context(self):
        """Add pipeline to serializer context"""
        context = super().get_serializer_context()
        if hasattr(self, 'get_object'):
            try:
                obj = self.get_object()
                context['pipeline'] = obj.pipeline
            except:
                pass
        return context
    
    @action(detail=True, methods=['post'])
    def soft_delete(self, request, pk=None):
        """Soft delete a record"""
        record = self.get_object()
        record.soft_delete(request.user)
        return Response({'message': 'Record deleted successfully'})
    
    @action(detail=True, methods=['post'])
    def restore(self, request, pk=None):
        """Restore a soft-deleted record"""
        record = get_object_or_404(Record, pk=pk, is_deleted=True)
        record.restore()
        serializer = self.get_serializer(record)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def process_ai_fields(self, request, pk=None):
        """Manually trigger AI field processing"""
        record = self.get_object()
        
        try:
            # COMMENTED OUT: Legacy manual AI processing - replace with new system
            # ai_results = process_ai_fields_sync(record, force_update=True)
            # if ai_results:
            #     record.data.update(ai_results)
            #     record.save(update_fields=['data'])
            
            # For now, trigger AI processing via the automatic system
            # by simulating a record save (which triggers _trigger_ai_updates)
            record.save()  # This triggers the modern AI system automatically
            
            serializer = self.get_serializer(record)
            return Response({
                'message': 'AI fields processing triggered via modern system',
                'record': serializer.data
            })
        
        except Exception as e:
            return Response(
                {'error': f'AI processing failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'])
    def bulk_action(self, request):
        """Perform bulk actions on records"""
        serializer = BulkRecordActionSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        data = serializer.validated_data
        action_type = data['action']
        record_ids = data['record_ids']
        
        # Get records user has access to
        records = self.get_queryset().filter(id__in=record_ids)
        
        if not records.exists():
            return Response(
                {'error': 'No accessible records found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        result = {'processed': 0, 'errors': []}
        
        if action_type == 'delete':
            for record in records:
                record.soft_delete(request.user)
                result['processed'] += 1
        
        elif action_type == 'update_status':
            new_status = data['status']
            records.update(status=new_status, updated_by=request.user)
            result['processed'] = records.count()
        
        elif action_type == 'add_tags':
            tags = data['tags']
            for record in records:
                record.tags = list(set(record.tags + tags))
                record.updated_by = request.user
                record.save(update_fields=['tags', 'updated_by'])
                result['processed'] += 1
        
        elif action_type == 'remove_tags':
            tags = data['tags']
            for record in records:
                record.tags = [tag for tag in record.tags if tag not in tags]
                record.updated_by = request.user
                record.save(update_fields=['tags', 'updated_by'])
                result['processed'] += 1
        
        elif action_type == 'export':
            # Return export data instead of performing action
            export_data = []
            for record in records:
                export_data.append(record.to_dict(include_metadata=True))
            
            return Response({
                'export_data': export_data,
                'count': len(export_data)
            })
        
        return Response(result)


class PipelineTemplateViewSet(viewsets.ModelViewSet):
    """ViewSet for managing pipeline templates"""
    queryset = PipelineTemplate.objects.all()
    serializer_class = PipelineTemplateSerializer
    permission_classes = [PipelinePermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'is_system', 'is_public']
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'usage_count', 'created_at']
    ordering = ['-usage_count', 'name']
    
    def get_queryset(self):
        """Filter templates based on visibility and permissions"""
        user = self.request.user
        if user.is_superuser:
            return PipelineTemplate.objects.all()
        
        # Use comprehensive permission system
        permission_manager = SyncPermissionManager(user)
        
        # Check if user has full template access
        if permission_manager.has_permission('action', 'pipelines', 'read_all'):
            return PipelineTemplate.objects.all()
        
        # Show public templates and user's own templates
        return PipelineTemplate.objects.filter(
            Q(is_public=True) |
            Q(created_by=user)
        )
    
    @action(detail=True, methods=['post'])
    def create_pipeline(self, request, pk=None):
        """Create a pipeline from this template"""
        template = self.get_object()
        serializer = PipelineTemplateCreatePipelineSerializer(
            data=request.data,
            context={'template': template, 'request': request}
        )
        
        if serializer.is_valid():
            pipeline = serializer.save()
            pipeline_serializer = PipelineSerializer(pipeline)
            return Response(pipeline_serializer.data, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['get'])
    def preview(self, request, pk=None):
        """Get template preview data"""
        template = self.get_object()
        
        return Response({
            'template_data': template.template_data,
            'preview_config': template.preview_config,
            'sample_data': template.sample_data
        })


class PipelineStatsViewSet(viewsets.ViewSet):
    """ViewSet for pipeline statistics"""
    permission_classes = [PipelinePermission]
    
    def list(self, request):
        """Get overall pipeline statistics"""
        user = request.user
        
        # Get user's accessible pipelines using permission system
        permission_manager = SyncPermissionManager(user)
        
        if permission_manager.has_permission('action', 'pipelines', 'read_all'):
            pipelines = Pipeline.objects.all()
        else:
            accessible_pipeline_ids = self._get_accessible_pipeline_ids(user, permission_manager)
            pipelines = Pipeline.objects.filter(id__in=accessible_pipeline_ids)
        
        stats = {
            'total_pipelines': pipelines.count(),
            'total_records': Record.objects.filter(
                pipeline__in=pipelines,
                is_deleted=False
            ).count(),
            'pipeline_types': {},
            'recent_activity': []
        }
        
        # Pipeline type breakdown
        for pipeline_type, _ in Pipeline.PIPELINE_TYPES:
            count = pipelines.filter(pipeline_type=pipeline_type).count()
            if count > 0:
                stats['pipeline_types'][pipeline_type] = count
        
        # Recent activity
        recent_records = Record.objects.filter(
            pipeline__in=pipelines,
            is_deleted=False
        ).select_related('pipeline').order_by('-updated_at')[:10]
        
        for record in recent_records:
            stats['recent_activity'].append({
                'id': record.id,
                'title': record.title,
                'pipeline': record.pipeline.name,
                'updated_at': record.updated_at
            })
        
        return Response(stats)
    
    def _get_accessible_pipeline_ids(self, user, permission_manager):
        """Get pipeline IDs that user has access to"""
        accessible_ids = []
        
        for pipeline in Pipeline.objects.all():
            if (permission_manager.has_permission('action', 'pipelines', 'read', str(pipeline.id)) or
                permission_manager.has_permission('action', 'pipelines', 'read') or
                pipeline.created_by == user or
                pipeline.access_level in ['public', 'internal']):
                accessible_ids.append(pipeline.id)
        
        return accessible_ids
    
    @action(detail=False, methods=['get'])
    def pipeline_stats(self, request):
        """Get detailed statistics for a specific pipeline"""
        pipeline_id = request.query_params.get('pipeline_id')
        if not pipeline_id:
            return Response(
                {'error': 'pipeline_id parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            pipeline = Pipeline.objects.get(id=pipeline_id)
        except Pipeline.DoesNotExist:
            return Response(
                {'error': 'Pipeline not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        records = pipeline.records.filter(is_deleted=False)
        
        stats = {
            'pipeline': {
                'id': pipeline.id,
                'name': pipeline.name,
                'type': pipeline.pipeline_type
            },
            'total_records': records.count(),
            'status_breakdown': {},
            'creation_timeline': [],
            'ai_field_stats': {}
        }
        
        # Status breakdown
        status_counts = records.values('status').annotate(
            count=models.Count('id')
        )
        for item in status_counts:
            stats['status_breakdown'][item['status']] = item['count']
        
        # AI field statistics
        ai_fields = pipeline.fields.filter(is_ai_field=True)
        for field in ai_fields:
            ai_records = records.exclude(data__isnull=True).exclude(
                **{f'data__{field.slug}__isnull': True}
            )
            stats['ai_field_stats'][field.name] = {
                'processed_records': ai_records.count(),
                'success_rate': (ai_records.count() / records.count() * 100) if records.count() > 0 else 0
            }
        
        return Response(stats)